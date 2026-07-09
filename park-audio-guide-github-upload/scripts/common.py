from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = Path.home() / "Desktop" / "全点位"
DATA_DIR = PROJECT_ROOT / "data"
PUBLIC_DIR = PROJECT_ROOT / "public"
OUTPUT_DIR = PROJECT_ROOT / "output"
CACHE_DIR = PROJECT_ROOT / "cache"
POINTS_JSON = DATA_DIR / "points.json"
PRONUNCIATION_JSON = DATA_DIR / "pronunciation_dict.json"
CONFIG_NAME = "点位配置.json"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
TEXT_EXTS = {".docx", ".txt", ".md"}
AUDIO_EXTS = {".mp3", ".m4a", ".wav", ".aac"}


def ensure_dirs() -> None:
    for path in [
        DATA_DIR,
        PUBLIC_DIR,
        PUBLIC_DIR / "p",
        PUBLIC_DIR / "assets" / "audio",
        PUBLIC_DIR / "assets" / "images",
        PUBLIC_DIR / "assets" / "css",
        PUBLIC_DIR / "assets" / "js",
        OUTPUT_DIR,
        OUTPUT_DIR / "qrcodes",
        OUTPUT_DIR / "samples",
        CACHE_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = PROJECT_ROOT / ".env"
    if env_path.exists():
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_order_and_name(folder_name: str) -> tuple[int | None, str]:
    name = folder_name.strip()
    match = re.match(r"^\s*(\d+)\s*[-_\.、 ]*\s*(.+?)\s*$", name)
    if match:
        return int(match.group(1)), match.group(2).strip()
    return None, name


def sort_key_for_folder(path: Path) -> tuple[int, str]:
    order, display_name = parse_order_and_name(path.name)
    return (order if order is not None else 999999, display_name)


def stable_short_key(folder_name: str) -> str:
    digest = hashlib.md5(folder_name.encode("utf-8")).hexdigest()[:6]
    return f"p-{digest}"


def safe_filename(text: str, suffix: str = "") -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "_", text.strip())
    cleaned = cleaned.strip("._") or "point"
    return cleaned + suffix


def select_text_file(files: list[Path]) -> Path | None:
    candidates = [p for p in files if p.suffix.lower() in TEXT_EXTS and not p.name.startswith("~$")]
    if not candidates:
        return None
    priority = {".docx": 0, ".txt": 1, ".md": 2}
    candidates.sort(key=lambda p: (priority.get(p.suffix.lower(), 9), len(p.name), p.name))
    return candidates[0]


def select_image_files(files: list[Path]) -> list[Path]:
    images = [p for p in files if p.suffix.lower() in IMAGE_EXTS and not p.name.startswith(".")]
    images.sort(key=lambda p: (0 if "图片" in p.stem or "cover" in p.stem.lower() else 1, p.name))
    return images


def select_audio_file(files: list[Path]) -> Path | None:
    preferred = [p for p in files if p.name == "AI讲解.mp3"]
    if preferred:
        return preferred[0]
    audios = [p for p in files if p.suffix.lower() in AUDIO_EXTS]
    audios.sort(key=lambda p: p.name)
    return audios[0] if audios else None


def read_text_file(path: Path | None) -> str:
    if path is None:
        return ""
    suffix = path.suffix.lower()
    if suffix == ".docx":
        try:
            from docx import Document

            document = Document(str(path))
            return "\n".join(p.text for p in document.paragraphs if p.text.strip())
        except Exception as exc:
            return f"[读取 docx 失败：{exc}]"
    if suffix in {".txt", ".md"}:
        for encoding in ("utf-8", "utf-8-sig", "gb18030"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_text(errors="ignore")
    return ""


def load_pronunciation_dict() -> dict[str, str]:
    if not PRONUNCIATION_JSON.exists():
        write_json(
            PRONUNCIATION_JSON,
            {
                "京博": "京博",
                "博华": "博华",
                "凌霄花": "凌霄花",
                "孔子草": "孔子草",
                "重组木": "重组木",
                "农文旅": "农文旅",
            },
        )
    return read_json(PRONUNCIATION_JSON, {})


def clean_text(raw: str) -> str:
    text = raw.replace("\r\n", "\n").replace("\r", "\n")
    replacements = {
        "，": "，",
        ",": "，",
        ";": "；",
        ":": "：",
        "?": "？",
        "!": "！",
        "(": "（",
        ")": "）",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    lines: list[str] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            lines.append("")
            continue
        if re.fullmatch(r"(页眉|页脚|批注|修订|第\s*\d+\s*页|Page\s*\d+).*", line, re.I):
            continue
        line = re.sub(r"^\s*[\d一二三四五六七八九十]+[\.、）)]\s*", "", line)
        line = re.sub(r"\s+", " ", line)
        lines.append(line)
    text = "\n".join(lines)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    for word, spoken in load_pronunciation_dict().items():
        if word and spoken:
            text = text.replace(word, spoken)
    return text


def text_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def tts_text(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"([。！？；：])", r"\1\n", text)
    text = re.sub(r"([，、])", r"\1 ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def split_for_tts(text: str, max_chars: int = 900) -> list[str]:
    text = tts_text(text)
    paragraphs = [p.strip() for p in re.split(r"\n{2,}", text) if p.strip()]
    chunks: list[str] = []
    current = ""
    for paragraph in paragraphs:
        pieces = re.split(r"(?<=[。！？；])", paragraph)
        for piece in [p.strip() for p in pieces if p.strip()]:
            if len(current) + len(piece) + 1 > max_chars and current:
                chunks.append(current.strip())
                current = piece
            else:
                current = f"{current}\n{piece}".strip()
        if current and len(current) > max_chars * 0.65:
            chunks.append(current.strip())
            current = ""
    if current:
        chunks.append(current.strip())
    return chunks or [text[:max_chars]]


def load_points(limit: int | None = None, point_key: str | None = None) -> list[dict[str, Any]]:
    points = read_json(POINTS_JSON, [])
    if point_key:
        points = [p for p in points if p.get("point_key") == point_key]
    points.sort(key=lambda p: (p.get("show_order") or 999999, p.get("display_name") or ""))
    if limit is not None:
        points = points[:limit]
    return points


def save_workbook(path: Path, headers: list[str], rows: list[list[Any]], red_flags: dict[tuple[int, int], bool] | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "清单"
    header_fill = PatternFill("solid", fgColor="EAF3EA")
    danger_fill = PatternFill("solid", fgColor="FCE4E4")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F3B2D")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    red_flags = red_flags or {}
    for (row_idx, col_idx), flagged in red_flags.items():
        if flagged:
            ws.cell(row=row_idx, column=col_idx).fill = danger_fill
            ws.cell(row=row_idx, column=col_idx).font = Font(color="B42318", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        values = [str(ws.cell(row=r, column=col_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(len(header), *(min(len(v), 60) for v in values)) + 3, 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def render_template(path: Path, context: dict[str, Any]) -> str:
    template = path.read_text(encoding="utf-8")
    for key, value in context.items():
        template = template.replace("{{ " + key + " }}", str(value))
    return template


def copy_file(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def run_checked(cmd: list[str]) -> tuple[bool, str]:
    try:
        result = subprocess.run(cmd, check=False, capture_output=True, text=True)
        return result.returncode == 0, (result.stdout + result.stderr).strip()
    except FileNotFoundError as exc:
        return False, str(exc)


def parse_common_args(description: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--limit", type=int, default=None, help="只处理前 N 个点位")
    parser.add_argument("--point", default=None, help="只处理指定 point_key")
    parser.add_argument("--force", action="store_true", help="强制重新生成")
    return parser


def public_audio_path(point_key: str) -> Path:
    return PUBLIC_DIR / "assets" / "audio" / f"{point_key}.mp3"


def page_url(point_key: str) -> str:
    return f"/p/{point_key}/"


def strip_number_for_check(display_name: str) -> str:
    return re.sub(r"^\s*\d+\s*[-_\.、 ]*", "", display_name).strip()


def html_escape(text: str) -> str:
    return html.escape(text or "", quote=True)


def paragraphs_to_html(text: str) -> str:
    parts = [p.strip() for p in re.split(r"\n{1,}", text or "") if p.strip()]
    return "\n".join(f"<p>{html_escape(p)}</p>" for p in parts)
